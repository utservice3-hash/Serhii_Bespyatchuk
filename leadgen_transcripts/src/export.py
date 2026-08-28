"""Deliverables: one text file per call, plus a spreadsheet index.

The text files are the readable artefact; the spreadsheet is the one-row-per-call
index that can be uploaded to Google Sheets.  Both carry the three things the
brief requires for every call: phone number, date, and the role-tagged text.
"""

from __future__ import annotations

import csv
import json
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from transcribe import Transcript

log = logging.getLogger(__name__)

COLUMNS = [
    "deal_id", "deal_name", "deal_closed_at", "manager",
    "phone", "call_date", "direction", "duration_sec",
    "role_confidence", "language", "recording_url", "transcript_file",
    "transcript",
]

_UNSAFE = re.compile(r"[^0-9A-Za-z._-]+")


@dataclass
class CallRecord:
    deal_id: int
    deal_name: str
    deal_closed_at: str
    manager: str
    phone: str
    call_date: str
    direction: str
    duration_sec: int
    recording_url: str
    transcript: Transcript
    transcript_file: str = ""

    def row(self) -> dict:
        return {
            "deal_id": self.deal_id,
            "deal_name": self.deal_name,
            "deal_closed_at": self.deal_closed_at,
            "manager": self.manager,
            "phone": self.phone,
            "call_date": self.call_date,
            "direction": {"in": "вхідний", "out": "вихідний"}.get(
                self.direction, self.direction),
            "duration_sec": self.duration_sec,
            "role_confidence": self.transcript.role_confidence,
            "language": self.transcript.language,
            "recording_url": self.recording_url,
            "transcript_file": self.transcript_file,
            "transcript": self.transcript.as_text(),
        }


def _safe(part: str, limit: int = 40) -> str:
    cleaned = _UNSAFE.sub("_", str(part)).strip("_")
    return cleaned[:limit] or "x"


def write_transcript_file(rec: CallRecord, root: Path) -> Path:
    """One .txt per call, named so it sorts by deal then date."""
    folder = root / "transcripts" / f"deal_{rec.deal_id}"
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{_safe(rec.call_date)}__{_safe(rec.phone)}.txt"

    header = [
        f"Угода:        {rec.deal_id} — {rec.deal_name}",
        f"Менеджер:     {rec.manager or '—'}",
        f"Телефон:      {rec.phone}",
        f"Дата дзвінка: {rec.call_date}",
        f"Напрямок:     {rec.row()['direction']}",
        f"Тривалість:   {rec.duration_sec} сек",
        f"Ролі:         {rec.transcript.role_confidence}"
        + (f" — {rec.transcript.note}" if rec.transcript.note else ""),
    ]
    body = rec.transcript.as_text() or "(порожня транскрипція)"
    path.write_text("\n".join(header) + "\n" + "-" * 60 + "\n" + body + "\n",
                    encoding="utf-8")
    rec.transcript_file = str(path.relative_to(root))
    return path


def write_csv(records: list[CallRecord], root: Path) -> Path:
    root.mkdir(parents=True, exist_ok=True)
    path = root / "calls_index.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=COLUMNS)
        writer.writeheader()
        for rec in records:
            writer.writerow(rec.row())
    log.info("wrote %s (%d rows)", path, len(records))
    return path


def write_xlsx(records: list[CallRecord], root: Path) -> Path | None:
    """Spreadsheet version of the index, if openpyxl is installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError:
        log.warning("openpyxl not installed - skipping .xlsx (CSV still written)")
        return None

    widths = {"deal_name": 30, "manager": 20, "phone": 18, "call_date": 20,
              "recording_url": 40, "transcript_file": 34, "transcript": 120}
    wb = Workbook()
    ws = wb.active
    ws.title = "calls"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for rec in records:
        row = rec.row()
        ws.append([row[c] for c in COLUMNS])

    transcript_col = COLUMNS.index("transcript") + 1
    for row in ws.iter_rows(min_row=2, min_col=transcript_col,
                            max_col=transcript_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = \
            widths.get(name, 14)

    root.mkdir(parents=True, exist_ok=True)
    path = root / "calls_index.xlsx"
    wb.save(path)
    log.info("wrote %s (%d rows)", path, len(records))
    return path


def write_summary(records: list[CallRecord], root: Path, stats: dict) -> Path:
    """Machine-readable run summary, handy for spot-checking coverage."""
    root.mkdir(parents=True, exist_ok=True)
    path = root / "run_summary.json"
    by_confidence: dict[str, int] = {}
    for rec in records:
        key = rec.transcript.role_confidence
        by_confidence[key] = by_confidence.get(key, 0) + 1
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        **stats,
        "calls_transcribed": len(records),
        "deals_with_calls": len({r.deal_id for r in records}),
        "role_confidence": by_confidence,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                    encoding="utf-8")
    return path
