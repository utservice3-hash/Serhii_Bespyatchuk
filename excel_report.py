import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 45)


def _fmt_ts(ts) -> str:
    if not ts:
        return ""
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""


def build_ad_report_excel(report: dict, days: int, trend: list[dict] | None = None) -> bytes:
    """
    Будує xlsx із листами:
    - "Зведення по кампаніях" — агреговані метрики по кожній рекламній кампанії
    - "Деталізація угод" — повний список угод, що прийшли з реклами за період
    - "Тренд (12 тижнів)" (якщо передано trend) — потижнева динаміка лідів/виручки
      за останні ~3 місяці з лінійним графіком
    """
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Зведення по кампаніях"
    summary_headers = [
        "Кампанія", "Всього угод", "Успіх", "Закрито не реалізовано",
        "Не цільові", "В роботі", "Конверсія %",
        "Витрати (оцінка), грн", "Виручка, грн", "Маржа, грн", "CPL, грн", "ROAS %",
    ]
    summary_ws.append(summary_headers)
    _style_header(summary_ws)

    campaigns = report.get("campaigns", {})
    has_spend = any(c.get("spend") is not None for c in campaigns.values())
    for name, c in sorted(campaigns.items(), key=lambda x: x[1]["total"], reverse=True):
        summary_ws.append([
            name, c["total"], c["won"], c["lost"], c["non_target"], c["in_progress"], c["conversion"],
            c.get("spend"), c.get("revenue"), c.get("margin"), c.get("cpl"), c.get("roas"),
        ])

    total_row = [
        "Всього", report.get("total", 0),
        sum(c["won"] for c in campaigns.values()),
        sum(c["lost"] for c in campaigns.values()),
        report.get("non_target_total", 0),
        sum(c["in_progress"] for c in campaigns.values()),
        "",
        sum(c["spend"] for c in campaigns.values() if c.get("spend") is not None) if has_spend else "",
        sum(c["revenue"] for c in campaigns.values()) if has_spend else "",
        sum(c["margin"] for c in campaigns.values() if c.get("margin") is not None) if has_spend else "",
        "",
        "",
    ]
    summary_ws.append(total_row)
    for cell in summary_ws[summary_ws.max_row]:
        cell.font = Font(bold=True)
    summary_ws.freeze_panes = "A2"
    _autosize(summary_ws)

    detail_ws = wb.create_sheet("Деталізація угод")
    detail_headers = [
        "ID угоди", "Назва угоди", "Кампанія", "Менеджер", "Статус",
        "Сума (грн)", "Дата створення", "Дата закриття",
    ]
    detail_ws.append(detail_headers)
    _style_header(detail_ws)

    for lead in report.get("leads", []):
        detail_ws.append([
            lead["lead_id"], lead["name"], lead["campaign"], lead["manager_name"],
            lead["status_label"], lead["amount"],
            _fmt_ts(lead["created_at"]), _fmt_ts(lead["closed_at"]),
        ])
    detail_ws.freeze_panes = "A2"
    detail_ws.auto_filter.ref = detail_ws.dimensions
    _autosize(detail_ws)

    if trend:
        trend_ws = wb.create_sheet("Тренд (12 тижнів)")
        trend_ws.append(["Тиждень з", "Всього угод", "Успіх", "Конверсія %", "Виручка, грн"])
        _style_header(trend_ws)
        for w in trend:
            trend_ws.append([w["week_start"], w["total"], w["won"], w["conversion"], w["revenue"]])
        _autosize(trend_ws)

        chart = LineChart()
        chart.title = "Динаміка лідів і виручки за 12 тижнів"
        chart.style = 12
        chart.y_axis.title = "Угоди / тис. грн"
        chart.x_axis.title = "Тиждень з"
        max_row = trend_ws.max_row

        leads_ref = Reference(trend_ws, min_col=2, max_col=2, min_row=1, max_row=max_row)
        chart.add_data(leads_ref, titles_from_data=True)
        revenue_ref = Reference(trend_ws, min_col=5, max_col=5, min_row=1, max_row=max_row)
        chart.add_data(revenue_ref, titles_from_data=True)
        cats = Reference(trend_ws, min_col=1, min_row=2, max_row=max_row)
        chart.set_categories(cats)
        chart.width, chart.height = 24, 12
        trend_ws.add_chart(chart, "G2")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
