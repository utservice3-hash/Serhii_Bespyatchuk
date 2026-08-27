import openpyxl,sys,json,datetime,re
from collections import defaultdict
wb=openpyxl.load_workbook(sys.argv[1],data_only=True)
out=defaultdict(dict)
for ws in wb.worksheets:
    t=ws.title
    if t.lower().startswith('dashboard') or t.lower().startswith('new dashboard') or t in ('Планувальник',): continue
    # verify layout
    if str(ws.cell(1,3).value or '').strip()!='Холодна база': continue
    for r in range(5,ws.max_row+1):
        v=ws.cell(r,1).value
        if not isinstance(v,datetime.datetime): continue
        g=lambda c: ws.cell(r,c).value if isinstance(ws.cell(r,c).value,(int,float)) else None
        d=v.strftime('%Y-%m-%d')
        goal=ws.cell(r,19).value
        out[t][d]=dict(cold_leads=g(3),cold_opr=g(4),cold_pro=g(5),
                       re_leads=g(7),re_opr=g(8),re_pro=g(9),
                       trucks=g(11),sum_trucks=g(12),sum_wait=g(13),
                       calls=g(15),calls_ok=g(16),viber=g(18),
                       goal=str(goal) if goal else '')
json.dump(out,open(sys.argv[2],'w'),ensure_ascii=False)
for t,dd in out.items():
    bym=defaultdict(list)
    for d,v in dd.items(): bym[d[:7]].append(v)
    print('#####',t)
    print(f"  {'міс':7s} {'акт.дн':>6s} {'спроб':>6s} {'успіш':>6s} {'%':>5s} {'спроб/д':>7s} {'ХБліди':>6s} {'ОПР':>5s} {'прорах':>6s} {'РЕліди':>6s} {'РЕопр':>5s} {'РЕпрор':>6s} {'маш':>4s} {'сума':>10s} {'вайбер':>6s}")
    for m in sorted(bym):
        v=bym[m]; s=lambda f: sum(x[f] or 0 for x in v)
        act=[x for x in v if any((x[f] or 0)>0 for f in ('calls','cold_leads','re_leads','viber','trucks'))]
        if not act: continue
        ca,ok=s('calls'),s('calls_ok')
        print(f"  {m} {len(act):6d} {ca:6.0f} {ok:6.0f} {(ok/ca*100 if ca else 0):5.1f} {(ca/len(act)):7.1f} {s('cold_leads'):6.0f} {s('cold_opr'):5.0f} {s('cold_pro'):6.0f} {s('re_leads'):6.0f} {s('re_opr'):5.0f} {s('re_pro'):6.0f} {s('trucks'):4.0f} {s('sum_trucks'):10.0f} {s('viber'):6.0f}")
